import openpyxl
import os

base = r'C:\Users\ruijie\Desktop\IDC数据文件'

# 1. Find actual data start in Forecast Product sheet
fpath = os.path.join(base, 'IDC China Quarterly Ethernet Switch Forecast, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
ws = wb['Product']
print('=== Switch Forecast 2025Q4 - Product (rows 1-20) ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
    non_null = [(j, str(c)[:20]) for j, c in enumerate(row) if c is not None]
    if non_null:
        print(f'  Row{i}: {non_null[:15]}')
wb.close()

# 2. WLAN Tracker - wlan sheet
fpath2 = os.path.join(base, 'IDC China Quarterly WLAN Tracker, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath2, read_only=True, data_only=True)
ws = wb['wlan']
print('\n=== WLAN Tracker 2025Q4 - wlan sheet ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
    vals = [str(c)[:25] if c is not None else '-' for c in row][:20]
    print(f'  Row{i}: {vals}')
wb.close()

# 3. Router Tracker - router sheet
fpath3 = os.path.join(base, 'IDC China Quarterly Router Tracker, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath3, read_only=True, data_only=True)
ws = wb['router']
print('\n=== Router Tracker 2025Q4 - router sheet ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
    vals = [str(c)[:25] if c is not None else '-' for c in row][:20]
    print(f'  Row{i}: {vals}')
wb.close()

# 4. WLAN Tracker-Segmentation - wlan sheet
fpath4 = os.path.join(base, 'IDC China Quarterly WLAN Tracker-Segmentation, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath4, read_only=True, data_only=True)
print(f'\n=== WLAN Tracker-Segmentation 2025Q4 - sheets: {wb.sheetnames} ===')
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if rows and rows[0] and rows[0][0] is not None:
        print(f'  [{sn}] Row1: {[str(c)[:25] if c is not None else "-" for c in rows[0]][:18]}')
wb.close()

# 5. Router Tracker-Segmentation - router sheet
fpath5 = os.path.join(base, 'IDC China Quarterly Router Tracker-Segmentation, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath5, read_only=True, data_only=True)
print(f'\n=== Router Tracker-Segmentation 2025Q4 - sheets: {wb.sheetnames} ===')
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if rows and rows[0] and rows[0][0] is not None:
        print(f'  [{sn}] Row1: {[str(c)[:25] if c is not None else "-" for c in rows[0]][:18]}')
wb.close()

# 6. WLAN Forecast - Product sheet (find data start)
fpath6 = os.path.join(base, 'IDC China Quarterly WLAN Forecast, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath6, read_only=True, data_only=True)
print(f'\n=== WLAN Forecast 2025Q4 - sheets: {wb.sheetnames} ===')
for sn in wb.sheetnames:
    ws = wb[sn]
    found = False
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        non_null = [(j, str(c)[:20]) for j, c in enumerate(row) if c is not None]
        if non_null:
            print(f'  [{sn}] Row{i}: {non_null[:10]}')
            found = True
            if i >= 3:
                break
    if not found:
        print(f'  [{sn}] - all empty in first 15 rows')
wb.close()

# 7. Forecast-Segmentation - Vertical sheet
fpath7 = os.path.join(base, 'IDC China Quarterly WLAN Forecast-Segmentation, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath7, read_only=True, data_only=True)
ws = wb['Vertical']
print('\n=== WLAN Forecast-Segmentation 2025Q4 - Vertical sheet (rows 1-15) ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
    non_null = [(j, str(c)[:20]) for j, c in enumerate(row) if c is not None]
    if non_null:
        print(f'  Row{i}: {non_null[:15]}')
wb.close()

# 8. Check existing output
fpath8 = os.path.join(base, '汇总结果', 'WLAN.xlsx')
if os.path.exists(fpath8):
    wb = openpyxl.load_workbook(fpath8, read_only=True, data_only=True)
    print(f'\n=== Existing output WLAN.xlsx - sheets: {wb.sheetnames} ===')
    for sn in wb.sheetnames[:3]:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
            non_null = [(j, str(c)[:25]) for j, c in enumerate(row) if c is not None]
            if non_null:
                print(f'  [{sn}] Row{i}: {non_null[:15]}')
    wb.close()
