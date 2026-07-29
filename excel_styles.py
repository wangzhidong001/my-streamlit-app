from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE7"),
    right=Side(style="thin", color="D9DEE7"),
    top=Side(style="thin", color="D9DEE7"),
    bottom=Side(style="thin", color="D9DEE7"),
)
ZEBRA_FILLS = [
    PatternFill(fill_type="solid", fgColor="FFFFFF"),
    PatternFill(fill_type="solid", fgColor="F2F7FB"),
]
CODE_FONT = Font(name="Consolas", size=10, color="2E75B6")
NAME_FONT = Font(name="Arial", size=10, bold=True, color="1F2937")
DESC_FONT = Font(name="Arial", size=10, color="374151")
CAT_FONT = Font(name="Arial", size=9, color="6B7280")


def apply_header(ws, headers, row):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def apply_data_rows(ws, data, start_row):
    for row_idx, (code, name, desc, cat) in enumerate(data, start_row):
        fill = ZEBRA_FILLS[(row_idx - start_row) % 2]
        ws.cell(row=row_idx, column=1, value=code).font = CODE_FONT
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=2, value=name).font = NAME_FONT
        ws.cell(row=row_idx, column=3, value=desc).font = DESC_FONT
        ws.cell(row=row_idx, column=4, value=cat).font = CAT_FONT
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in range(1, 5):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 48
