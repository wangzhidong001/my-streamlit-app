"""
IDC 数据处理流水线
步骤1-7: 文件扫描 → 历史对比 → 分类 → 实际数汇总 → 预测数汇总 → 全产品汇总 → 分析计算
"""
import os
import re
import json
import shutil
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook

# ============ 配置 ============
# 数据目录自动适配：
# 1. 优先读取环境变量 DATA_DIR（Streamlit Cloud 等云端部署）
# 2. 本地运行时若桌面数据文件夹存在，沿用旧路径保证兼容性
# 3. 否则使用仓库内的 ./data 目录（云端默认、本地也可直接用）
_LOCAL_DESKTOP_DIR = r'C:\Users\ruijie\Desktop\IDC数据文件'
if os.environ.get('DATA_DIR'):
    DATA_DIR = os.environ.get('DATA_DIR')
elif os.path.exists(_LOCAL_DESKTOP_DIR):
    DATA_DIR = _LOCAL_DESKTOP_DIR
else:
    DATA_DIR = './data'

OUTPUT_DIR = os.environ.get('OUTPUT_DIR', os.path.join(DATA_DIR, '汇总结果'))
BACKUP_DIR = os.path.join(OUTPUT_DIR, '历史版本备份')
HISTORY_FILE = os.path.join(OUTPUT_DIR, '处理历史记录.json')
SOURCE_HISTORY_FILE = os.path.join(DATA_DIR, '处理历史.json')

PRODUCTS = ['WLAN', 'Switch', 'Router', 'VCC']
PRODUCT_SHEET_MAP = {
    'Switch': 'switch',
    'WLAN': 'wlan',
    'Router': 'router',
    'VCC': 'VCC'
}
PRODUCT_KEYWORDS = {
    'Switch': ['Ethernet Switch'],
    'WLAN': ['WLAN'],
    'Router': ['Router'],
    'VCC': ['Virtual Client Computing']
}

# 字段更名映射
COLUMN_RENAME = {
    'Layer': 'Technology',
    'Ports': 'Units',
    'Place in Network': '二级产品分类',
    'Connectivity': '二级产品分类',
    'Location': '二级产品分类',
    'Speed': '末级产品分类',
    'Standard': '末级产品分类',
    'Segment': '行业大类',
    'Information': 'Deployment',
    'Place.in.Network': 'Place in Network',
    # Tracker 文件中的列名更名
    'Company': 'Vendor',
    'Value (CNY¥M)': 'Vendor Revenue (CNY M)',
    'Value (US$M)': 'Vendor Revenue (USD M)',
    'Value (CNY M)': 'Vendor Revenue (CNY M)',
    'Value (US M)': 'Vendor Revenue (USD M)',
    'Value (USD M)': 'Vendor Revenue (USD M)',
    'Vendor Revenue (US$M)': 'Vendor Revenue (USD M)',
    'Vendor Revenue (US M)': 'Vendor Revenue (USD M)',
}

# Step 7 最终输出列顺序
FINAL_COLUMNS = [
    '产品/行业', 'Year', '实际/预测', '预测Year', '预测版本',
    'Half Year', 'Quarter', 'Technology',
    '二级产品分类', 'Product', 'Product Detail', '末级产品分类',
    'Deployment', '行业大类', '行业细分',
    'Vendor', 'Units', 'Vendor Revenue (CNY M)', 'Vendor Revenue (USD M)'
]


# ============ 步骤1: 扫描文件 ============
def scan_excel_files():
    """扫描IDC数据文件文件夹，只获取顶层xlsx文件"""
    files = []
    for f in os.listdir(DATA_DIR):
        if f.startswith('~$'):
            continue
        if f.endswith('.xlsx') or f.endswith('.xls'):
            fpath = os.path.join(DATA_DIR, f)
            if os.path.isfile(fpath):
                files.append({
                    'filename': f,
                    'filepath': fpath,
                    'mtime': os.path.getmtime(fpath)
                })
    return files


# ============ 步骤2: 历史对比 ============
def load_history():
    """加载处理历史"""
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    if os.path.exists(SOURCE_HISTORY_FILE):
        with open(SOURCE_HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'last_processed': None, 'processed_files': []}


def save_history(history):
    """保存处理历史"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def check_new_files(files, history):
    """对比历史记录，识别新增或修改的文件"""
    processed_map = {}
    for pf in history.get('processed_files', []):
        processed_map[pf['filename']] = pf.get('mtime', 0)
    
    new_files = []
    all_files = []
    for f in files:
        all_files.append(f['filename'])
        old_mtime = processed_map.get(f['filename'])
        if old_mtime is None or f['mtime'] > old_mtime:
            new_files.append(f)
    
    return new_files, all_files


# ============ 步骤4: 文件分类 ============
def classify_file(filename):
    """
    判断数据文件类型
    返回: dict(product, data_type, perspective, period, version)
    """
    info = {
        'product': None,
        'data_type': None,       # 'actual' / 'forecast'
        'perspective': None,     # 'product' / 'industry'
        'period': None,          # '2025Q2', '2025H2' etc
        'version': None,         # 'Q2', 'H2' etc
        'year': None             # 2025
    }
    
    # 判断产品
    for product, keywords in PRODUCT_KEYWORDS.items():
        if any(kw.lower() in filename.lower() for kw in keywords):
            info['product'] = product
            break
    
    # 判断数据类型
    if 'forecast' in filename.lower():
        info['data_type'] = 'forecast'
    elif 'tracker' in filename.lower():
        info['data_type'] = 'actual'
    
    # 判断口径
    if 'segmentation' in filename.lower():
        info['perspective'] = 'industry'
    else:
        info['perspective'] = 'product'
    
    # 提取时间版本 (如 2025Q2, 2025H2)
    period_match = re.search(r',\s*(\d{4}[QH]\d)', filename)
    if period_match:
        info['period'] = period_match.group(1)
        info['year'] = int(info['period'][:4])
        info['version'] = info['period'][-2:]
    
    return info


def classify_all_files(files):
    """分类所有文件"""
    classified = {}
    for f in files:
        info = classify_file(f['filename'])
        info['filepath'] = f['filepath']
        info['filename'] = f['filename']
        info['mtime'] = f['mtime']
        
        key = f"{info['product']}_{info['data_type']}_{info['perspective']}_{info['period']}"
        classified[key] = info
    
    return classified


def get_latest_files(classified):
    """获取每个时间段的最新版文件"""
    # 按 (product, data_type, perspective) 分组，取最新period
    groups = {}
    for key, info in classified.items():
        group_key = f"{info['product']}_{info['data_type']}_{info['perspective']}"
        if group_key not in groups:
            groups[group_key] = []
        groups[group_key].append(info)
    
    latest = {}
    for group_key, items in groups.items():
        items.sort(key=lambda x: x['period'] or '', reverse=True)
        latest[group_key] = items[0]
    
    return latest


# ============ 步骤5: 实际数汇总 ============
def clean_currency(val):
    """清除货币符号"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    for sym in ['￥', '$', '¥', '€', '£', ',']:
        s = s.replace(sym, '')
    s = s.strip()
    if s == '' or s == 'nan' or s == 'None':
        return None
    try:
        return float(s)
    except ValueError:
        return s


def rename_columns(df, product):
    """字段更名"""
    df = df.copy()
    
    # VCC特殊处理
    if product == 'VCC':
        if 'Vendor Revenue' in df.columns:
            df = df.rename(columns={'Vendor Revenue': 'Vendor Revenue (USD M)'})
        if 'Segment' in df.columns:
            df = df.rename(columns={'Segment': '行业大类'})
        if 'Information' in df.columns:
            df = df.rename(columns={'Information': 'Deployment'})
        return df
    
    # 处理重复列：2026Q1文件同时有Company/Vendor, Value/Vendor Revenue等
    # 如果目标列已存在，删除源列；否则重命名
    drop_cols = []
    conflict_renames = {
        'Company': 'Vendor',
        'Value (CNY¥M)': 'Vendor Revenue (CNY M)',
        'Value (US$M)': 'Vendor Revenue (USD M)',
        'Value (CNY M)': 'Vendor Revenue (CNY M)',
        'Value (US M)': 'Vendor Revenue (USD M)',
        'Value (USD M)': 'Vendor Revenue (USD M)',
        'Vendor Revenue (US$M)': 'Vendor Revenue (USD M)',
    }
    for src, dst in conflict_renames.items():
        if src in df.columns and dst in df.columns:
            drop_cols.append(src)
    if drop_cols:
        df = df.drop(columns=drop_cols)
    
    # 通用更名：Place.in.Network → Place in Network
    rename_map = {}
    for col in df.columns:
        col_str = str(col).strip()
        if col_str == 'Place.in.Network':
            rename_map[col] = 'Place in Network'
    if rename_map:
        df = df.rename(columns=rename_map)
    
    # 第二轮更名：使用 COLUMN_RENAME 字典
    rename_map2 = {}
    for col in df.columns:
        col_str = str(col).strip()
        if col_str in COLUMN_RENAME and COLUMN_RENAME[col_str] != col_str:
            rename_map2[col] = COLUMN_RENAME[col_str]
    if rename_map2:
        df = df.rename(columns=rename_map2)
    
    # 处理重命名后可能产生的重复列名：保留第一个
    if df.columns.duplicated().any():
        dup_cols = df.columns[df.columns.duplicated()].tolist()
        print(f"  ⚠ 检测到重复列名: {dup_cols}，保留第一个")
        df = df.loc[:, ~df.columns.duplicated()]
    
    return df


def fill_technology(df, product):
    """填充Technology字段"""
    if 'Technology' not in df.columns:
        df['Technology'] = product
    else:
        df['Technology'] = df['Technology'].fillna(product)
        # Ethernet Switch -> Switch
        df['Technology'] = df['Technology'].replace({'Ethernet Switch': 'Switch'})
        df['Technology'] = df['Technology'].fillna(product)
    return df


def process_quarter_field(df):
    """Quarter字段只保留后两个字符"""
    if 'Quarter' in df.columns:
        df['Quarter'] = df['Quarter'].astype(str).str[-2:]
    return df


def ensure_half_year(df):
    """如果没有Half Year字段，根据Year和Quarter转换"""
    if 'Half Year' not in df.columns or df['Half Year'].isna().all():
        if 'Year' in df.columns and 'Quarter' in df.columns:
            def convert_hy(row):
                year = str(row.get('Year', ''))
                q = str(row.get('Quarter', ''))
                if q.startswith('Q1') or q.startswith('Q2'):
                    return f'{year}H1'
                elif q.startswith('Q3') or q.startswith('Q4'):
                    return f'{year}H2'
                return None
            df['Half Year'] = df.apply(convert_hy, axis=1)
    return df


def read_tracker_sheet(filepath, product):
    """读取Tracker文件的对应sheet"""
    sheet_name = PRODUCT_SHEET_MAP.get(product)
    df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
    return df


def merge_perspectives(product_df, industry_df, product):
    """
    合并产品口径和行业口径数据
    以产品口径字段顺序为基础，行业口径特有字段插在Vendor Revenue (CNY M)前面
    """
    if product_df is None and industry_df is None:
        return None
    if product_df is None:
        return industry_df
    if industry_df is None:
        return product_df
    
    # 获取产品口径的列顺序
    base_cols = list(product_df.columns)
    ind_cols = list(industry_df.columns)
    
    # 找出行业口径特有的列
    extra_cols = [c for c in ind_cols if c not in base_cols]
    
    # 找到插入位置：Vendor Revenue (CNY M) 前面
    insert_pos = len(base_cols)
    for i, c in enumerate(base_cols):
        if 'Vendor Revenue (CNY M)' in str(c):
            insert_pos = i
            break
    
    # 构建合并后的列顺序
    merged_cols = base_cols[:insert_pos] + extra_cols + base_cols[insert_pos:]
    
    # 确保两个df都有所有列
    for c in merged_cols:
        if c not in product_df.columns:
            product_df[c] = None
        if c not in industry_df.columns:
            industry_df[c] = None
    
    merged = pd.concat([product_df[merged_cols], industry_df[merged_cols]], ignore_index=True)
    return merged


def process_actual_data(classified):
    """步骤5: 汇总实际数"""
    print("\n=== 步骤5: 汇总实际数 ===")
    
    # 获取最新的实际数文件
    latest = get_latest_files(classified)
    
    product_results = {}
    
    for product in PRODUCTS:
        # 获取产品口径实际数
        prod_key = f"{product}_actual_product"
        ind_key = f"{product}_actual_industry"
        
        prod_info = latest.get(prod_key)
        ind_info = latest.get(ind_key)
        
        if product == 'VCC':
            # VCC只有一种口径
            if prod_info is None:
                # VCC可能没有perspective分类
                for key, info in classified.items():
                    if info['product'] == 'VCC' and info['data_type'] == 'actual':
                        prod_info = info
                        break
            
            if prod_info is None:
                print(f"  {product}: 无实际数文件")
                continue
            
            df = read_tracker_sheet(prod_info['filepath'], product)
            df = rename_columns(df, product)
            df = fill_technology(df, product)
            
            # 添加产品/行业字段
            df['产品/行业'] = '产品口径'
            df['实际/预测'] = '实际'
            
            # 清理数值列
            numeric_cols = [c for c in df.columns if 'Revenue' in str(c) or 'Value' in str(c)]
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = df[col].apply(clean_currency)
            
            product_results[product] = df
            print(f"  {product}: {len(df)} 行实际数")
            continue
        
        # 其他产品：读取产品口径和行业口径
        prod_df = None
        ind_df = None
        
        if prod_info:
            prod_df = read_tracker_sheet(prod_info['filepath'], product)
            prod_df = rename_columns(prod_df, product)
            prod_df = fill_technology(prod_df, product)
            prod_df = process_quarter_field(prod_df)
            prod_df = ensure_half_year(prod_df)
            prod_df['产品/行业'] = '产品口径'
            prod_df['实际/预测'] = '实际'
            print(f"  {product} 产品口径: {len(prod_df)} 行")
        
        if ind_info:
            ind_df = read_tracker_sheet(ind_info['filepath'], product)
            ind_df = rename_columns(ind_df, product)
            ind_df = fill_technology(ind_df, product)
            ind_df = process_quarter_field(ind_df)
            ind_df = ensure_half_year(ind_df)
            ind_df['产品/行业'] = '行业口径'
            ind_df['实际/预测'] = '实际'
            print(f"  {product} 行业口径: {len(ind_df)} 行")
        
        if prod_df is None and ind_df is None:
            print(f"  {product}: 无实际数文件")
            continue
        
        # 合并两个口径
        merged = merge_perspectives(prod_df, ind_df, product)
        
        # 清理数值列
        numeric_cols = [c for c in merged.columns if 'Revenue' in str(c) or 'Value' in str(c)]
        for col in numeric_cols:
            if col in merged.columns:
                merged[col] = merged[col].apply(clean_currency)
        
        # 验证：同一年的行业口径数 = 产品口径数
        verify_actual_data(merged, product)
        
        product_results[product] = merged
        print(f"  {product} 合并后: {len(merged)} 行")
    
    return product_results


def verify_actual_data(df, product):
    """验证：同一年的行业口径数等于产品口径数"""
    if '产品/行业' not in df.columns or 'Year' not in df.columns:
        return
    if 'Vendor Revenue (CNY M)' not in df.columns:
        return
    
    for year in df['Year'].dropna().unique():
        year_data = df[df['Year'] == year]
        prod_sum = year_data[year_data['产品/行业'] == '产品口径']['Vendor Revenue (CNY M)'].sum()
        ind_sum = year_data[year_data['产品/行业'] == '行业口径']['Vendor Revenue (CNY M)'].sum()
        if prod_sum > 0 and ind_sum > 0:
            diff = abs(prod_sum - ind_sum) / max(prod_sum, ind_sum) * 100
            if diff > 1:
                print(f"  ⚠ {product} {year}年: 产品口径={prod_sum:.2f} vs 行业口径={ind_sum:.2f} 差异={diff:.2f}%")


# ============ 步骤6: 预测数汇总 ============
def read_forecast_pivot(filepath, sheet_name):
    """读取Forecast透视表，返回行转列后的DataFrame"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # 找到包含多个年份数字的行作为表头行（如2024, 2025, 2026...）
    header_row_idx = None
    for i, row in enumerate(rows):
        if row:
            year_count = sum(1 for c in row if c is not None and str(c).strip().isdigit() and 2020 <= int(str(c).strip()) <= 2035)
            if year_count >= 3:
                header_row_idx = i
                break
    
    if header_row_idx is None:
        print(f"    ⚠ 未找到表头行: {sheet_name}")
        return None
    
    header_row = rows[header_row_idx]
    data_rows = rows[header_row_idx + 1:]
    
    # 跳过空的首列，找到第一个非空列
    start_col = 0
    for j, c in enumerate(header_row):
        if c is not None and str(c).strip() != '':
            start_col = j
            break
    
    header_row = header_row[start_col:]
    data_rows = [row[start_col:] for row in data_rows if row is not None]
    
    # 构建列名
    col_names = []
    for i, c in enumerate(header_row):
        if c is None or str(c).strip() == '':
            col_names.append(f'col_{i}')
        else:
            col_names.append(str(c).strip())
    
    # 构建DataFrame
    df = pd.DataFrame(data_rows, columns=col_names)
    
    # 过滤掉全空的行
    df = df.dropna(how='all')
    
    # 前向填充非年份列
    non_year_cols = [c for c in df.columns if not str(c).isdigit()]
    df[non_year_cols] = df[non_year_cols].ffill()
    
    # 过滤掉Total行和Grand Total行
    def is_total_row(row_vals):
        for v in row_vals:
            v_str = str(v).lower() if v is not None else ''
            if 'total' in v_str or 'grand' in v_str:
                return True
        return False
    
    df = df[~df.apply(lambda r: is_total_row(r.values), axis=1)]
    
    # 行转列：年份列转为行
    year_cols = [c for c in df.columns if str(c).isdigit()]
    if not year_cols:
        return df
    
    id_vars = [c for c in df.columns if c not in year_cols]
    melted = df.melt(id_vars=id_vars, value_vars=year_cols, var_name='Year', value_name='Vendor Revenue (CNY M)')
    melted = melted.dropna(subset=['Vendor Revenue (CNY M)'])
    melted['Vendor Revenue (CNY M)'] = melted['Vendor Revenue (CNY M)'].apply(clean_currency)
    melted = melted[melted['Vendor Revenue (CNY M)'] != 0]
    melted = melted[melted['Vendor Revenue (CNY M)'].notna()]
    
    print(f"    {sheet_name}: 表头在第{header_row_idx+1}行, {len(melted)}行数据, 年份列={year_cols}")
    return melted


def process_forecast_data(classified, actual_results):
    """步骤6: 汇总预测数 - 保留所有预测版本"""
    print("\n=== 步骤6: 汇总预测数 ===")
    
    forecast_results = {}
    
    for product in PRODUCTS:
        actual_df = actual_results.get(product)
        if actual_df is None:
            print(f"  {product}: 无实际数基础，跳过预测")
            continue
        
        all_forecast_dfs = []
        
        if product == 'VCC':
            # VCC预测数从Tracker文件的VCC Forecast data sheet获取
            for key, info in classified.items():
                if info['product'] == 'VCC' and info['data_type'] == 'actual':
                    fc_df = read_vcc_forecast(info['filepath'], info['period'])
                    if fc_df is not None and len(fc_df) > 0:
                        fc_df['产品/行业'] = '产品口径'
                        all_forecast_dfs.append(fc_df)
                        print(f"  VCC 预测数: {len(fc_df)} 行 (版本: {info.get('period', 'N/A')})")
        else:
            # 遍历所有预测版本（保留每个版本）
            for key, info in classified.items():
                if info['product'] != product or info['data_type'] != 'forecast':
                    continue
                
                period = info.get('period', '')
                pred_year = period[:4] if period else None
                pred_ver = period[-2:] if period else None
                
                if info['perspective'] == 'industry':
                    # 行业口径预测：从Segmentation Forecast的Vertical sheet获取
                    fc_df = read_forecast_pivot(info['filepath'], 'Vertical')
                    if fc_df is not None and len(fc_df) > 0:
                        fc_df = rename_forecast_columns(fc_df, product, 'industry')
                        fc_df['产品/行业'] = '行业口径'
                        fc_df['预测Year'] = pred_year
                        fc_df['预测版本'] = pred_ver
                        all_forecast_dfs.append(fc_df)
                        print(f"  {product} 行业口径预测: {len(fc_df)} 行 (版本: {period})")
                
                elif info['perspective'] == 'product':
                    # 产品口径预测
                    if product in ('WLAN', 'Router'):
                        fc_df = read_forecast_pivot(info['filepath'], 'Product')
                    elif product == 'Switch':
                        fc_df = read_forecast_pivot(info['filepath'], 'DC Product')
                    else:
                        fc_df = None
                    
                    if fc_df is not None and len(fc_df) > 0:
                        fc_df = rename_forecast_columns(fc_df, product, 'product')
                        fc_df['产品/行业'] = '产品口径'
                        fc_df['预测Year'] = pred_year
                        fc_df['预测版本'] = pred_ver
                        all_forecast_dfs.append(fc_df)
                        print(f"  {product} 产品口径预测: {len(fc_df)} 行 (版本: {period})")
        
        if not all_forecast_dfs:
            print(f"  {product}: 无预测数文件")
            forecast_results[product] = actual_df.copy()
            continue
        
        # 合并所有预测版本
        forecast_combined = pd.concat(all_forecast_dfs, ignore_index=True)
        forecast_combined['实际/预测'] = '预测'
        
        # 对齐实际数的列结构
        actual_cols = list(actual_df.columns)
        for c in actual_cols:
            if c not in forecast_combined.columns:
                forecast_combined[c] = None
        # 只取实际数中有的列，保留预测数特有列
        common_cols = [c for c in actual_cols if c in forecast_combined.columns]
        extra_cols = [c for c in forecast_combined.columns if c not in actual_cols]
        forecast_combined = forecast_combined[common_cols + extra_cols]
        
        # 合并实际数和预测数
        combined = pd.concat([actual_df, forecast_combined], ignore_index=True)
        
        # 排序：行业口径在上、产品口径在下；实际在上、预测在下
        combined['_sort_persp'] = combined['产品/行业'].map({'行业口径': 0, '产品口径': 1})
        combined['_sort_type'] = combined['实际/预测'].map({'实际': 0, '预测': 1})
        combined = combined.sort_values(['_sort_persp', '_sort_type', 'Year', '预测版本']).drop(columns=['_sort_persp', '_sort_type'])
        
        forecast_results[product] = combined
        print(f"  {product} 合并后: {len(combined)} 行 (实际+预测)")
    
    return forecast_results


def read_vcc_forecast(filepath, period):
    """读取VCC Forecast data sheet"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    if 'VCC Forecast data' not in wb.sheetnames:
        wb.close()
        return None
    ws = wb['VCC Forecast data']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # 找到Year行和Revenue行
    records = []
    year_row = None
    half_year_row = None
    revenue_row = None
    
    for i, row in enumerate(rows):
        if row and any(str(c).strip() == 'Year' for c in row if c is not None):
            year_row = row
        if row and any(str(c).strip() == 'Half Year' for c in row if c is not None):
            half_year_row = row
        if row and any(str(c).strip() == 'Revenue' for c in row if c is not None):
            revenue_row = row
    
    if year_row is None or revenue_row is None:
        return None
    
    # 构建数据
    for i in range(1, len(year_row)):
        year_val = year_row[i]
        rev_val = revenue_row[i] if i < len(revenue_row) else None
        if year_val and rev_val and year_val != 'Year':
            records.append({
                'Year': str(year_val),
                'Vendor Revenue (USD M)': clean_currency(rev_val),
                '实际/预测': '预测',
                'Technology': 'VCC'
            })
    
    df = pd.DataFrame(records)
    if len(df) > 0:
        pred_year = period[:4] if period else None
        pred_ver = period[-2:] if period else None
        df['预测Year'] = pred_year
        df['预测版本'] = pred_ver
    
    return df


def rename_forecast_columns(df, product, perspective):
    """重命名预测数列以匹配实际数结构"""
    col_map = {}
    for c in df.columns:
        c_str = str(c).strip()
        # 使用 COLUMN_RENAME 字典统一处理
        if c_str in COLUMN_RENAME:
            col_map[c] = COLUMN_RENAME[c_str]
        # Standard, Location 也需要更名（WLAN Product sheet）
        elif c_str == 'Standard':
            col_map[c] = '末级产品分类'
        elif c_str == 'Location':
            col_map[c] = '二级产品分类'
    
    if col_map:
        df = df.rename(columns=col_map)
    
    # 填充Technology
    if 'Technology' not in df.columns:
        df['Technology'] = 'Switch' if product == 'Switch' else product
    else:
        df['Technology'] = df['Technology'].fillna(product)
        df['Technology'] = df['Technology'].replace({'Ethernet Switch': 'Switch'})
    
    return df


# ============ 步骤7: 全产品汇总 ============
def combine_all_products(forecast_results):
    """步骤7: 汇总四个产品数据"""
    print("\n=== 步骤7: 全产品汇总 ===")
    
    all_dfs = []
    for product in PRODUCTS:
        df = forecast_results.get(product)
        if df is None:
            continue
        
        # 确保所有最终列都存在
        for col in FINAL_COLUMNS:
            if col not in df.columns:
                df[col] = None
        
        all_dfs.append(df[FINAL_COLUMNS])
        print(f"  {product}: {len(df)} 行")
    
    if not all_dfs:
        print("  无数据可汇总")
        return None
    
    combined = pd.concat(all_dfs, ignore_index=True)
    
    # 排序
    combined['_sort_persp'] = combined['产品/行业'].map({'行业口径': 0, '产品口径': 1})
    combined['_sort_type'] = combined['实际/预测'].map({'实际': 0, '预测': 1})
    combined = combined.sort_values(['_sort_persp', '_sort_type', 'Technology', 'Year'])
    combined = combined.drop(columns=['_sort_persp', '_sort_type'])
    
    print(f"  全产品汇总: {len(combined)} 行")
    return combined.copy()


def calculate_analysis(combined_df, vat_rate=0.13,
                       industry_filter=None,
                       perspective_filter=None,
                       dc_category=None,
                       vendor_filter=None,
                       deployment_filter=None):
    """步骤7.3: 分析计算
    
    Args:
        combined_df: 全产品汇总数据
        vat_rate: 增值税率
        industry_filter: 行业大类筛选列表，None或空列表=不筛选
        perspective_filter: 产品/行业口径筛选列表，None或空列表=不筛选
        dc_category: DC产品分类筛选列表，None或空列表=不筛选
        vendor_filter: Vendor厂商筛选列表，None或空列表=不筛选
        deployment_filter: Deployment字段筛选列表，None或空列表=不筛选
    """
    def _norm(v):
        """规范化筛选值为列表，None/空→空列表"""
        if v is None:
            return []
        if isinstance(v, (list, tuple)):
            return [x for x in v if x]
        return [v]
    
    industry_list = _norm(industry_filter)
    perspective_list = _norm(perspective_filter)
    dc_list = _norm(dc_category)
    vendor_list = _norm(vendor_filter)
    deployment_list = _norm(deployment_filter)
    
    print("\n=== 步骤7.3: 分析计算 ===")
    print(f"  筛选条件: 行业={industry_list or '全部'}, 口径={perspective_list or '全部'}, DC分类={dc_list or '全部'}, Vendor={vendor_list or '全部'}, Deployment={deployment_list or '全部'}")
    
    results = []
    
    # 获取实际数最大年
    actual_data = combined_df[combined_df['实际/预测'] == '实际']
    if len(actual_data) == 0:
        print("  无实际数数据")
        return pd.DataFrame()
    
    # ===== 3.1 实际数加工 =====
    print("  --- 实际数加工 ---")
    
    def _apply_filters(df):
        """逐个应用筛选条件，空列表跳过"""
        mask = pd.Series([True] * len(df), index=df.index)
        if perspective_list:
            mask &= df['产品/行业'].isin(perspective_list)
        if industry_list:
            mask &= df['行业大类'].isin(industry_list)
        if deployment_list:
            mask &= df['Deployment'].isin(deployment_list)
        if dc_list:
            mask &= df['二级产品分类'].isin(dc_list)
        if vendor_list:
            mask &= df['Vendor'].apply(
                lambda v: any(str(s).lower() in str(v).lower() for s in vendor_list) if pd.notna(v) else False
            )
        return df[mask]
    
    # 1. 获取行业全产品实际数（分两步：先按口径+行业筛选，DC和Vendor在汇总时再应用）
    def _apply_outer_filters(df):
        """应用口径、行业、Deployment筛选（外层筛选）"""
        mask = pd.Series([True] * len(df), index=df.index)
        if perspective_list:
            mask &= df['产品/行业'].isin(perspective_list)
        if industry_list:
            mask &= df['行业大类'].isin(industry_list)
        if deployment_list:
            mask &= df['Deployment'].isin(deployment_list)
        return df[mask]
    
    telecom_actual = _apply_outer_filters(actual_data)
    
    # 回退：如果无结果且有Deployment条件，尝试Deployment筛选
    if len(telecom_actual) == 0 and deployment_list:
        telecom_actual = actual_data[actual_data['Deployment'].isin(deployment_list)]
    
    # 最终回退：如果仍无结果且无任何外层筛选，使用全部实际数
    if len(telecom_actual) == 0 and not perspective_list and not industry_list and not deployment_list:
        telecom_actual = actual_data
    
    # 基于通信行业实际数确定预测起始年份和预测版本
    if len(telecom_actual) > 0:
        telecom_max_year = telecom_actual['Year'].dropna().astype(str).max()
        telecom_max_year_data = telecom_actual[telecom_actual['Year'] == telecom_max_year]
        telecom_quarters = telecom_max_year_data['Quarter'].dropna().unique()
        telecom_has_q4 = any('Q4' in str(q) for q in telecom_quarters)
        
        if telecom_has_q4:
            forecast_start_year = int(telecom_max_year) + 1
        else:
            forecast_start_year = int(telecom_max_year)
        
        # 确定应该使用的预测版本
        # 找到最大的预测版本
        forecast_data_all = combined_df[combined_df['实际/预测'] == '预测']
        forecast_versions = sorted([v for v in forecast_data_all['预测版本'].dropna().unique() if str(v) != 'nan'])
        forecast_years_ver = sorted([y for y in forecast_data_all['预测Year'].dropna().unique() if str(y) != 'nan'], reverse=True)
        latest_forecast_version = None
        if forecast_years_ver:
            latest_forecast_year = forecast_years_ver[0]
            latest_versions = sorted(forecast_data_all[forecast_data_all['预测Year'] == latest_forecast_year]['预测版本'].dropna().unique(), reverse=True)
            if latest_versions:
                latest_forecast_version = (latest_forecast_year, latest_versions[0])
        
        print(f"  通信行业实际数最大年: {telecom_max_year}, Q4: {telecom_has_q4}")
        print(f"  预测起始年份: {forecast_start_year}")
        print(f"  最新预测版本: {latest_forecast_version}")
    else:
        forecast_start_year = None
        latest_forecast_version = None
    
    if len(telecom_actual) > 0:
        # 按Year汇总
        for year in sorted(telecom_actual['Year'].dropna().unique()):
            year_data = telecom_actual[telecom_actual['Year'] == year]
            total_revenue = year_data['Vendor Revenue (CNY M)'].sum()
            
            # 检查Quarter
            year_quarters = year_data['Quarter'].dropna().unique()
            has_full_year = any('Q4' in str(q) for q in year_quarters)
            display_year = str(year)
            if not has_full_year:
                display_year = f'{year}H'
            
            # 2. 获取DC容量
            if dc_list:
                dc_data = year_data[year_data['二级产品分类'].isin(dc_list)]
            else:
                dc_data = year_data
            dc_revenue = dc_data['Vendor Revenue (CNY M)'].sum()
            
            # 3. DC占全产品比例
            dc_ratio = dc_revenue / total_revenue * 100 if total_revenue > 0 else 0
            
            results.append({
                '数据类型': '实际',
                '年份': display_year,
                '通信全产品容量': round(total_revenue, 2),
                '通信DC容量': round(dc_revenue, 2),
                'DC占全产品比例': round(dc_ratio, 2),
            })
    
    # 4-9: 计算增速、锐捷份额等
    actual_results = [r for r in results if r['数据类型'] == '实际']
    for i, r in enumerate(actual_results):
        year_str = str(r['年份']).replace('H', '')
        year_int = int(year_str)
        
        # 4. DC容量增速
        if i > 0:
            prev_dc = actual_results[i-1]['通信DC容量']
            if prev_dc > 0:
                r['通信DC容量增速'] = round((r['通信DC容量'] - prev_dc) / prev_dc * 100, 2)
            else:
                r['通信DC容量增速'] = None
        else:
            r['通信DC容量增速'] = None
        
        # 5. 目标Vendor DC收入
        year_telecom = telecom_actual[telecom_actual['Year'] == year_str]
        if dc_list:
            dc_mask = year_telecom['二级产品分类'].isin(dc_list)
        else:
            dc_mask = pd.Series([True] * len(year_telecom), index=year_telecom.index)
        
        if vendor_list:
            vendor_mask = year_telecom['Vendor'].apply(
                lambda v: any(str(s).lower() in str(v).lower() for s in vendor_list) if pd.notna(v) else False
            )
        else:
            vendor_mask = pd.Series([True] * len(year_telecom), index=year_telecom.index)
        
        dc_ruijie = year_telecom[dc_mask & vendor_mask]
        ruijie_dc_revenue = dc_ruijie['Vendor Revenue (CNY M)'].sum()
        r['锐捷DC收入'] = round(ruijie_dc_revenue, 2)
        
        # 6. 锐捷DC份额
        if r['通信DC容量'] > 0:
            r['锐捷DC份额'] = round(ruijie_dc_revenue / r['通信DC容量'] * 100, 2)
        else:
            r['锐捷DC份额'] = 0
        
        # 7. 竞争力指数
        if i > 0 and actual_results[i-1].get('锐捷DC份额', 0) > 0:
            r['竞争力指数'] = round(r['锐捷DC份额'] / actual_results[i-1]['锐捷DC份额'], 2)
        else:
            r['竞争力指数'] = None
        
        # 8. 开票金额
        r['增值税率'] = round(vat_rate, 2)
        r['锐捷开票金额'] = round(ruijie_dc_revenue * (1 + vat_rate), 2)
        
        # 9. 开票同比变动
        if i > 0 and actual_results[i-1].get('锐捷开票金额', 0) > 0:
            prev_inv = actual_results[i-1]['锐捷开票金额']
            r['开票同比变动'] = round((r['锐捷开票金额'] - prev_inv) / prev_inv * 100, 2)
        else:
            r['开票同比变动'] = None

    # ===== 3.2 预测数加工 =====
    print("  --- 预测数加工 ---")
    
    forecast_data = combined_df[combined_df['实际/预测'] == '预测']
    if len(forecast_data) > 0 and len(actual_results) > 0 and forecast_start_year is not None:
        
        # 对每个产品(Technology)+口径组合使用其最新的预测版本
        print(f"  对每个产品+口径使用最新预测版本:")
        latest_forecast_dfs = []
        for tech in forecast_data['Technology'].dropna().unique():
            for persp in forecast_data['产品/行业'].dropna().unique():
                subset = forecast_data[
                    (forecast_data['Technology'] == tech) &
                    (forecast_data['产品/行业'] == persp)
                ]
                if len(subset) == 0:
                    continue
                # 找到该组合最新的预测版本
                versions = []
                for _, row in subset[['预测Year', '预测版本']].drop_duplicates().iterrows():
                    py = str(row['预测Year']).strip()
                    pv = str(row['预测版本']).strip()
                    if py != 'nan' and pv != 'nan':
                        versions.append((py, pv))
                if versions:
                    versions.sort(reverse=True)
                    latest_ver = versions[0]
                    latest_data = subset[
                        (subset['预测Year'].astype(str).str.strip() == latest_ver[0]) &
                        (subset['预测版本'].astype(str).str.strip() == latest_ver[1])
                    ]
                    latest_forecast_dfs.append(latest_data)
                    print(f"    {tech} {persp}: 最新版本 {latest_ver[0]}{latest_ver[1]}, {len(latest_data)} 行")
        
        if latest_forecast_dfs:
            forecast_data = pd.concat(latest_forecast_dfs, ignore_index=True)
            print(f"  合并后预测数据: {len(forecast_data)} 行")
        
        # 获取最新DC比例
        last_actual = actual_results[-1] if actual_results else {}
        last_dc_ratio = last_actual.get('DC占全产品比例', 0) / 100
        
        # 1. 获取行业全产品预测数
        telecom_forecast = _apply_outer_filters(forecast_data)
        
        # 回退：如果无结果且有Deployment条件
        if len(telecom_forecast) == 0 and deployment_list:
            telecom_forecast = forecast_data[forecast_data['Deployment'].isin(deployment_list)]
        
        # 最终回退
        if len(telecom_forecast) == 0 and not perspective_list and not industry_list and not deployment_list:
            telecom_forecast = forecast_data
        
        if len(telecom_forecast) > 0:
            # 获取所有符合条件的预测年份
            all_forecast_years = sorted([y for y in telecom_forecast['Year'].dropna().unique() 
                                       if str(y).strip().isdigit() and int(str(y).strip()) >= forecast_start_year])
            
            forecast_years = all_forecast_years
            
            for i, year in enumerate(forecast_years):
                year_data = telecom_forecast[telecom_forecast['Year'] == year]
                total_revenue = year_data['Vendor Revenue (CNY M)'].sum()
                
                # 3. 通信DC容量 = 全产品容量 * DC比例
                dc_revenue = total_revenue * last_dc_ratio
                
                # 4. DC容量增速
                if i == 0 and actual_results:
                    prev_dc = actual_results[-1].get('通信DC容量', 0)
                elif i > 0:
                    prev_dc = results[len(actual_results) + i - 1].get('通信DC容量', 0) if len(results) > len(actual_results) + i - 1 else 0
                else:
                    prev_dc = 0
                
                dc_growth = round((dc_revenue - prev_dc) / prev_dc * 100, 2) if prev_dc > 0 else None
                
                result = {
                    '数据类型': '预测',
                    '年份': str(year),
                    '通信全产品容量': round(total_revenue, 2),
                    '通信DC容量': round(dc_revenue, 2),
                    'DC占全产品比例': round(last_dc_ratio * 100, 2),
                    '通信DC容量增速': dc_growth,
                }
                
                # 5. 锐捷DC份额 - 需要导入，这里暂用上一年
                if i == 0 and actual_results:
                    result['锐捷DC份额'] = actual_results[-1].get('锐捷DC份额', 0)
                elif i > 0 and len(results) > len(actual_results):
                    result['锐捷DC份额'] = results[len(actual_results) + i - 1].get('锐捷DC份额', 0)
                else:
                    result['锐捷DC份额'] = 0
                
                # 6. 锐捷DC收入
                ruijie_dc_revenue = dc_revenue * result['锐捷DC份额'] / 100
                result['锐捷DC收入'] = round(ruijie_dc_revenue, 2)
                
                # 7. 竞争力指数
                if i == 0 and actual_results and actual_results[-1].get('锐捷DC份额', 0) > 0:
                    result['竞争力指数'] = round(result['锐捷DC份额'] / actual_results[-1]['锐捷DC份额'], 2)
                elif i > 0 and len(results) > len(actual_results):
                    prev_share = results[len(actual_results) + i - 1].get('锐捷DC份额', 0)
                    result['竞争力指数'] = round(result['锐捷DC份额'] / prev_share, 2) if prev_share > 0 else None
                else:
                    result['竞争力指数'] = None
                
                # 8. 开票金额
                result['增值税率'] = round(vat_rate, 2)
                result['锐捷开票金额'] = round(ruijie_dc_revenue * (1 + vat_rate), 2)
                
                # 9. 开票同比变动
                if i == 0 and actual_results and actual_results[-1].get('锐捷开票金额', 0) > 0:
                    prev_inv = actual_results[-1]['锐捷开票金额']
                    result['开票同比变动'] = round((result['锐捷开票金额'] - prev_inv) / prev_inv * 100, 2)
                elif i > 0 and len(results) > len(actual_results):
                    prev_inv = results[len(actual_results) + i - 1].get('锐捷开票金额', 0)
                    result['开票同比变动'] = round((result['锐捷开票金额'] - prev_inv) / prev_inv * 100, 2) if prev_inv > 0 else None
                else:
                    result['开票同比变动'] = None

                results.append(result)
                print(f"    预测 {year}年: 全产品={total_revenue:.2f} DC={dc_revenue:.2f}")
    
    # 整理结果列顺序
    result_columns = [
        '数据类型', '年份', '通信全产品容量', '通信DC容量', 'DC占全产品比例',
        '通信DC容量增速', '锐捷DC收入', '锐捷DC份额', '竞争力指数',
        '增值税率', '锐捷开票金额'
    ]
    
    result_df = pd.DataFrame(results)
    for col in result_columns:
        if col not in result_df.columns:
            result_df[col] = None
    result_df = result_df[result_columns]
    
    print(f"  分析结果: {len(result_df)} 行")
    return result_df


def calculate_analysis_fixed(combined_df, vat_rate=0.13, share_csv=None):
    """固定版通信DC分析计算（复刻 comm_processor.py 第一版逻辑）

    固定筛选：
        - 产品/行业 = 行业口径
        - 行业大类   = 通信
        - DC 口径    = 二级产品分类 = Datacenter
        - 锐捷       = Vendor = Ruijie

    预测版本选择：
        - 实际最大年 Quarter 为 Q1~Q3 时，取该季度版本、起始年=最大年
        - 为 Q4 时，取 Q4 版本、起始年=最大年+1

    输出列（11列）：
        数据类型、年份、通信全产品容量、通信DC容量、DC占全产品比例、
        通信DC容量增速、锐捷DC收入、锐捷DC份额、竞争力指数、增值税率、锐捷开票金额
    """
    PRODUCT_LINE = '行业口径'
    INDUSTRY = '通信'
    DC_CLASS = 'Datacenter'
    VENDOR_RUIJIE = 'Ruijie'
    OUT_COLS = [
        '数据类型', '年份', '通信全产品容量', '通信DC容量', 'DC占全产品比例',
        '通信DC容量增速', '锐捷DC收入', '锐捷DC份额', '竞争力指数', '增值税率', '锐捷开票金额'
    ]

    def _pct(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        return round(x * 100, 2)

    def _annual_sum(df):
        return df.groupby('Year')['Vendor Revenue (CNY M)'].sum()

    base = combined_df[(combined_df['产品/行业'] == PRODUCT_LINE) &
                       (combined_df['行业大类'] == INDUSTRY)].copy()
    # 统一 Year 列为数值类型，避免字符串/整数混用导致 groupby 后 .get() 匹配失败
    base['Year'] = pd.to_numeric(base['Year'], errors='coerce')
    act = base[base['实际/预测'] == '实际'].copy()
    fc = base[base['实际/预测'] == '预测'].copy()

    if act.empty:
        print("  [错误] 无「行业口径/通信/实际」数据")
        return pd.DataFrame(columns=OUT_COLS)

    # 预测版本/起始年
    act_years = sorted(act['Year'].dropna().astype(int).unique())
    ymax = max(act_years)
    qmax_values = act[act['Year'] == ymax]['Quarter'].dropna().astype(str).unique()
    qmax = max(qmax_values) if len(qmax_values) > 0 else 'Q4'
    if qmax in ('Q1', 'Q2', 'Q3'):
        fver, start_year = qmax, ymax
    else:
        fver, start_year = 'Q4', ymax + 1
    print(f"  实际最大年={ymax}({qmax}) -> 预测版本={fver}, 起始年={start_year}")

    # ---- 3.1 实际数 ----
    years = sorted(act['Year'].dropna().astype(int).unique())
    qs = act.groupby('Year')['Quarter'].apply(lambda s: set(s.dropna().astype(str).unique()))
    label = {y: (str(y) if 'Q4' in qs.get(y, set()) else f"{y}H") for y in years}

    total = _annual_sum(act)
    dc = _annual_sum(act[act['二级产品分类'] == DC_CLASS])
    ruijie = _annual_sum(act[(act['二级产品分类'] == DC_CLASS) &
                             (act['Vendor'] == VENDOR_RUIJIE)])

    recs = {}
    for y in years:
        t = float(total.get(y, 0.0))
        d = float(dc.get(y, 0.0))
        r = float(ruijie.get(y, 0.0))
        ratio = (d / t) if t else None
        share = (r / d) if d else None
        recs[y] = dict(全产品=t, DC=d, ratio=ratio, ruijie=r, share=share)

    rows = []
    prev_dc = prev_share = prev_invoice = None
    for y in years:
        rec = recs[y]
        dc_growth = ((rec['DC'] - prev_dc) / prev_dc) if prev_dc not in (None, 0) else None
        comp_idx = (rec['share'] / prev_share) if prev_share not in (None, 0) else None
        invoice = rec['ruijie'] * (1 + vat_rate)
        invoice_yoy = ((invoice - prev_invoice) / prev_invoice) if prev_invoice not in (None, 0) else None
        rows.append({
            '数据类型': '实际',
            '年份': label[y],
            '通信全产品容量': round(rec['全产品'], 2),
            '通信DC容量': round(rec['DC'], 2),
            'DC占全产品比例': _pct(rec['ratio']),
            '通信DC容量增速': _pct(dc_growth),
            '锐捷DC收入': round(rec['ruijie'], 2),
            '锐捷DC份额': _pct(rec['share']),
            '竞争力指数': None if comp_idx is None else round(comp_idx, 4),
            '增值税率': round(vat_rate * 100, 2),
            '锐捷开票金额': round(invoice, 2),
            '_sort': y,
        })
        prev_dc, prev_share, prev_invoice = rec['DC'], rec['share'], invoice

    # ---- 3.2 预测数 ----
    share_input = None
    if share_csv and os.path.exists(share_csv):
        sd = pd.read_csv(share_csv)
        col_val = sd.columns[1]
        share_input = {}
        for _, r in sd.iterrows():
            y = int(r[sd.columns[0]])
            v = float(r[col_val])
            share_input[y] = v / 100.0 if v > 1 else v
        print(f"  已加载外部份额文件: {share_csv}")

    fc_sel = fc[(fc['预测版本'].astype(str) == fver) &
                (fc['Year'].astype(int) >= start_year)]
    total_fc = _annual_sum(fc_sel)
    ratio_fc = recs[ymax]['ratio']
    share_fallback = recs[ymax]['share']

    fc_years = sorted(total_fc.index.dropna().astype(int).unique())
    frecs = {}
    for y in fc_years:
        t = float(total_fc.get(y, 0.0))
        d = (t * ratio_fc) if ratio_fc is not None else None
        share_y = (share_input.get(y, share_fallback) if share_input else share_fallback)
        r = (d * share_y) if (d is not None and share_y is not None) else None
        frecs[y] = dict(全产品=t, DC=d, share=share_y, ruijie=r)

    prev_dc = recs[ymax]['DC']
    prev_share = recs[ymax]['share']
    prev_invoice = recs[ymax]['ruijie'] * (1 + vat_rate)
    for y in fc_years:
        rec = frecs[y]
        dc_growth = ((rec['DC'] - prev_dc) / prev_dc) if prev_dc not in (None, 0) else None
        comp_idx = (rec['share'] / prev_share) if prev_share not in (None, 0) else None
        invoice = rec['ruijie'] * (1 + vat_rate) if rec['ruijie'] is not None else None
        invoice_yoy = ((invoice - prev_invoice) / prev_invoice) if prev_invoice not in (None, 0) else None
        rows.append({
            '数据类型': '预测',
            '年份': str(y),
            '通信全产品容量': round(rec['全产品'], 2),
            '通信DC容量': None if rec['DC'] is None else round(rec['DC'], 2),
            'DC占全产品比例': _pct(ratio_fc),
            '通信DC容量增速': _pct(dc_growth),
            '锐捷DC收入': None if rec['ruijie'] is None else round(rec['ruijie'], 2),
            '锐捷DC份额': _pct(rec['share']),
            '竞争力指数': None if comp_idx is None else round(comp_idx, 4),
            '增值税率': round(vat_rate * 100, 2),
            '锐捷开票金额': None if invoice is None else round(invoice, 2),
            '_sort': y,
        })
        prev_dc, prev_share, prev_invoice = rec['DC'], rec['share'], invoice

    rows.sort(key=lambda r: (r['_sort'], 0 if r['数据类型'] == '实际' else 1))
    for r in rows:
        r.pop('_sort', None)

    result_df = pd.DataFrame(rows, columns=OUT_COLS)
    print(f"  固定版分析结果: {len(result_df)} 行")
    return result_df


def generate_chart_data(analysis_df):
    """生成图表数据"""
    if len(analysis_df) == 0:
        return None

    years = analysis_df['年份'].tolist()
    dc_capacity = analysis_df['通信DC容量'].tolist()
    ruijie_share = analysis_df['锐捷DC份额'].tolist()
    # 取锐捷DC收入（不再取锐捷DC容量）
    ruijie_dc = analysis_df['锐捷DC收入'].tolist()

    return {
        'years': years,
        'dc_capacity': dc_capacity,
        'ruijie_dc': ruijie_dc,
        'ruijie_share': ruijie_share
    }


# ============ 输出文件 ============
def save_product_files(forecast_results, output_dir):
    """保存各产品汇总文件"""
    os.makedirs(output_dir, exist_ok=True)
    for product, df in forecast_results.items():
        if df is None or len(df) == 0:
            continue
        fpath = os.path.join(output_dir, f'{product}.xlsx')
        df.to_excel(fpath, sheet_name=product, index=False)
        print(f"  保存: {product}.xlsx ({len(df)} 行)")


def save_combined_file(combined_df, output_dir):
    """保存全产品汇总文件"""
    if combined_df is None:
        return
    fpath = os.path.join(output_dir, 'IDC全产品数据.xlsx')
    combined_df.to_excel(fpath, sheet_name='IDC全产品数据', index=False)
    print(f"  保存: IDC全产品数据.xlsx ({len(combined_df)} 行)")


def save_analysis_file(analysis_df, output_dir, suffix=''):
    """保存分析结果
    
    Args:
        analysis_df: 分析结果数据
        output_dir: 输出目录
        suffix: 场景后缀，为空时保存为默认文件名
    """
    if len(analysis_df) == 0:
        return
    filename = f'IDC分析结果{f"_{suffix}" if suffix else ""}.xlsx'
    fpath = os.path.join(output_dir, filename)
    analysis_df.to_excel(fpath, sheet_name='分析结果', index=False)
    print(f"  保存: {filename} ({len(analysis_df)} 行)")


def save_chart_html(chart_data, output_dir, suffix=''):
    """保存图表HTML
    
    Args:
        chart_data: 图表数据
        output_dir: 输出目录
        suffix: 场景后缀，为空时保存为默认文件名
    """
    if chart_data is None:
        return
    
    import json as json_mod
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>通信DC容量与锐捷DC份额</title>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <style>
        body {{ margin: 0; padding: 20px; font-family: "Microsoft YaHei", sans-serif; }}
        h2 {{ text-align: center; color: #333; }}
        #chart {{ width: 100%; height: 500px; }}
    </style>
</head>
<body>
    <h2>通信DC容量与锐捷DC份额趋势</h2>
    <div id="chart"></div>
    <script>
        var data = {json_mod.dumps(chart_data, ensure_ascii=False)};
        var chart = echarts.init(document.getElementById('chart'));
        chart.setOption({{
            tooltip: {{ trigger: 'axis', axisPointer: {{ type: 'cross' }} }},
            legend: {{ data: ['通信DC容量', '锐捷DC收入', '锐捷DC份额'] }},
            grid: {{ left: '3%', right: '4%', bottom: '3%', containLabel: true }},
            xAxis: {{ type: 'category', data: data.years }},
            yAxis: [
                {{ type: 'value', name: '容量(CNY M)', position: 'left' }},
                {{ type: 'value', name: '份额(%)', position: 'right' }}
            ],
            series: [
                {{
                    name: '通信DC容量',
                    type: 'bar',
                    data: data.dc_capacity,
                    itemStyle: {{ color: '#5470c6' }}
                }},
                {{
                    name: '锐捷DC收入',
                    type: 'bar',
                    data: data.ruijie_dc,
                    itemStyle: {{ color: '#91cc75' }}
                }},
                {{
                    name: '锐捷DC份额',
                    type: 'line',
                    yAxisIndex: 1,
                    data: data.ruijie_share,
                    itemStyle: {{ color: '#ee6666' }},
                    lineStyle: {{ width: 3 }}
                }}
            ]
        }});
        window.addEventListener('resize', function() {{ chart.resize(); }});
    </script>
</body>
</html>'''
    
    fpath = os.path.join(output_dir, f'IDC图表{f"_{suffix}" if suffix else ""}.html')
    with open(fpath, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  保存: IDC图表{f'_{suffix}' if suffix else ''}.html")


# ============ 主流程 ============
def main():
    print("=" * 60)
    print("IDC 数据处理流水线")
    print("=" * 60)
    
    # 步骤1: 扫描文件
    print("\n=== 步骤1: 扫描文件 ===")
    files = scan_excel_files()
    print(f"  扫描到 {len(files)} 个Excel文件")
    
    # 步骤2: 历史对比
    print("\n=== 步骤2: 历史对比 ===")
    history = load_history()
    new_files, all_filenames = check_new_files(files, history)
    
    if not new_files:
        print("  无新增文件，使用上次处理结果")
        # 返回上次结果
        if os.path.exists(OUTPUT_DIR):
            print(f"  上次处理结果在: {OUTPUT_DIR}")
            return
    else:
        print(f"  发现 {len(new_files)} 个新增/修改文件:")
        for f in new_files:
            print(f"    - {f['filename']}")
    
    # 步骤3: 判断是否首次处理
    is_first = len(history.get('processed_files', [])) == 0
    print(f"\n=== 步骤3: {'首次处理' if is_first else '增量处理'} ===")
    
    # 备份旧结果
    if not is_first and os.path.exists(OUTPUT_DIR):
        backup_dir = os.path.join(BACKUP_DIR, datetime.now().strftime('%Y%m%d_%H%M%S'))
        os.makedirs(backup_dir, exist_ok=True)
        for f in os.listdir(OUTPUT_DIR):
            if f.endswith('.xlsx'):
                shutil.copy2(os.path.join(OUTPUT_DIR, f), os.path.join(backup_dir, f))
        print(f"  旧结果已备份到: {backup_dir}")
    
    # 步骤4: 文件分类
    print("\n=== 步骤4: 文件分类 ===")
    classified = classify_all_files(files)
    for key, info in classified.items():
        print(f"  {info['product']:8s} | {info['data_type']:8s} | {info['perspective']:10s} | {info['period'] or 'N/A':8s} | {os.path.basename(info['filename'])}")
    
    # 步骤5: 实际数汇总
    actual_results = process_actual_data(classified)
    
    # 步骤6: 预测数汇总
    forecast_results = process_forecast_data(classified, actual_results)
    
    # 步骤7: 全产品汇总
    combined = combine_all_products(forecast_results)
    
    # 分析计算
    analysis = calculate_analysis(combined)
    
    # 生成图表数据
    chart_data = generate_chart_data(analysis)
    
    # 保存结果
    print("\n=== 保存结果 ===")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    save_product_files(forecast_results, OUTPUT_DIR)
    save_combined_file(combined, OUTPUT_DIR)
    save_analysis_file(analysis, OUTPUT_DIR)
    save_chart_html(chart_data, OUTPUT_DIR)
    
    # 更新历史记录
    new_history = {
        'last_processed': datetime.now().isoformat(),
        'processed_files': [{'filename': f['filename'], 'mtime': f['mtime']} for f in files]
    }
    save_history(new_history)
    print(f"\n处理历史已更新: {HISTORY_FILE}")
    
    print("\n" + "=" * 60)
    print("处理完成！")
    print(f"结果输出目录: {OUTPUT_DIR}")
    print("=" * 60)


if __name__ == '__main__':
    main()
