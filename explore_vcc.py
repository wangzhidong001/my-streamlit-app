import openpyxl
import os

out_dir = r'C:\Users\ruijie\Desktop\IDC数据文件\汇总结果'

for fname in ['VCC.xlsx', 'Switch.xlsx', 'Router.xlsx']:
    fpath = os.path.join(out_dir, fname)
    if not os.path.exists(fpath):
        print(f'{fname} NOT FOUND')
        continue
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if row1 and row1[0]:
            print(f'=== {fname} [{sn}] ===')
            for i, c in enumerate(row1[0]):
                print(f'  Col{i}: {c}')
            # Find a few sample rows
            for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), 2):
                non_null = [(j, str(c)[:20]) for j, c in enumerate(row) if c is not None]
                if non_null:
                    print(f'  Row{ri}: {non_null[:15]}')
            print(f'  Max row: {ws.max_row}, Max col: {ws.max_column}')
            print()
    wb.close()

# Also check IDC全产品数据_new.xlsx
fpath2 = os.path.join(out_dir, 'IDC全产品数据_new.xlsx')
if os.path.exists(fpath2):
    wb = openpyxl.load_workbook(fpath2, read_only=True, data_only=True)
    print(f'=== IDC全产品数据_new.xlsx - sheets: {wb.sheetnames} ===')
    for sn in wb.sheetnames:
        ws = wb[sn]
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if row1 and row1[0]:
            print(f'  [{sn}] Header: {[str(c)[:25] if c else "-" for c in row1[0]][:25]}')
        print(f'  [{sn}] Max row: {ws.max_row}')
    wb.close()
