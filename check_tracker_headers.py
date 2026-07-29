"""检查Tracker文件表头结构"""
from openpyxl import load_workbook
import os

DATA_DIR = r'C:\Users\ruijie\Desktop\IDC数据文件'

files = [
    ('IDC China Quarterly WLAN Tracker, 2025Q4.xlsx', 'wlan'),
    ('IDC China Quarterly Ethernet Switch Tracker, 2025Q4.xlsx', 'switch'),
    ('IDC China Quarterly Router Tracker, 2025Q4.xlsx', 'router'),
]

for fname, sheet in files:
    fpath = os.path.join(DATA_DIR, fname)
    if not os.path.exists(fpath):
        print(f"NOT FOUND: {fname}")
        continue
    print(f"\n{'='*60}")
    print(f"{fname} - Sheet: {sheet}")
    wb = load_workbook(fpath, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"  Sheet '{sheet}' not found! Available: {wb.sheetnames}")
        wb.close()
        continue
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        vals = [str(c)[:25] if c is not None else '-' for c in row[:20]]
        print(f"  Row{i}: {vals}")
    wb.close()

# Also check VCC
vcc_file = 'IDC_China Semiannual Virtual Client Computing Software Tracker, 2025H2.xlsx'
fpath = os.path.join(DATA_DIR, vcc_file)
if os.path.exists(fpath):
    print(f"\n{'='*60}")
    print(f"{vcc_file}")
    wb = load_workbook(fpath, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  --- {sn} (max_row={ws.max_row}, max_col={ws.max_column}) ---")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            vals = [str(c)[:25] if c is not None else '-' for c in row[:20]]
            print(f"  Row{i}: {vals}")
    wb.close()
