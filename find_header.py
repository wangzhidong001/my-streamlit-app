"""找到预测数文件的实际数据起始行"""
from openpyxl import load_workbook
import os

DATA_DIR = r'C:\Users\ruijie\Desktop\IDC数据文件'

files_to_check = [
    ('IDC China Quarterly WLAN Forecast, 2025Q4.xlsx', 'Product'),
    ('IDC China Quarterly WLAN Forecast-Segmentation, 2025Q4.xlsx', 'Vertical'),
    ('IDC China Quarterly Ethernet Switch Forecast, 2025Q4.xlsx', 'DC Product'),
]

for fname, sheet_name in files_to_check:
    fpath = os.path.join(DATA_DIR, fname)
    if not os.path.exists(fpath):
        print(f"NOT FOUND: {fname}")
        continue
    
    print(f"\n{'='*70}")
    print(f"文件: {fname} | Sheet: {sheet_name}")
    print(f"{'='*70}")
    
    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    # 扫描前50行，找非空行
    found_data = False
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        # 检查是否有非空值
        non_empty = [c for c in row if c is not None and str(c).strip() != '']
        if non_empty:
            vals = [str(c)[:25] if c is not None else '-' for c in row[:27]]
            print(f"  Row{i}: {vals}")
            found_data = True
    
    if not found_data:
        # 尝试更多行
        print("  前50行全空，继续扫描...")
        for i, row in enumerate(ws.iter_rows(min_row=50, max_row=100, values_only=True), 50):
            non_empty = [c for c in row if c is not None and str(c).strip() != '']
            if non_empty:
                vals = [str(c)[:25] if c is not None else '-' for c in row[:27]]
                print(f"  Row{i}: {vals}")
                found_data = True
                if i > 70:
                    break
    
    wb.close()
