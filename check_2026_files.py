"""检查2026Q1文件和Segmentation文件的列名"""
from openpyxl import load_workbook
import os

DATA_DIR = r'C:\Users\ruijie\Desktop\IDC数据文件'

# Check 2026Q1 files
files = [
    ('IDC China Quarterly WLAN Tracker, 2026Q1.xlsx', 'wlan'),
    ('IDC China Quarterly Ethernet Switch Tracker, 2026Q1.xlsx', 'switch'),
    ('IDC China Quarterly Router Tracker, 2026Q1.xlsx', 'router'),
]

for fname, sheet in files:
    fpath = os.path.join(DATA_DIR, fname)
    if not os.path.exists(fpath):
        print(f"NOT FOUND: {fname}")
        continue
    print(f"\n{'='*60}")
    print(f"{fname} - Sheet: {sheet}")
    wb = load_workbook(fpath, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"  Sheets: {wb.sheetnames}")
        wb.close()
        continue
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
        vals = [str(c)[:30] if c is not None else '-' for c in row[:20]]
        print(f"  Row{i}: {vals}")
    wb.close()

# Check Segmentation files (Tracker)
seg_files = [
    'IDC China Quarterly WLAN Tracker-Segmentation, 2025Q4.xlsx',
    'IDC China Quarterly Ethernet Switch Tracker-Segmentation, 2025Q4.xlsx',
]

for fname in seg_files:
    fpath = os.path.join(DATA_DIR, fname)
    if not os.path.exists(fpath):
        print(f"NOT FOUND: {fname}")
        continue
    print(f"\n{'='*60}")
    print(f"{fname}")
    wb = load_workbook(fpath, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row > 1:
            print(f"\n  --- {sn} (rows={ws.max_row}) ---")
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
                vals = [str(c)[:30] if c is not None else '-' for c in row[:20]]
                print(f"  Row{i}: {vals}")
    wb.close()

# Check VCC Forecast data sheet
vcc_file = 'IDC_China Semiannual Virtual Client Computing Software Tracker, 2025H2.xlsx'
fpath = os.path.join(DATA_DIR, vcc_file)
if os.path.exists(fpath):
    print(f"\n{'='*60}")
    print(f"{vcc_file} - VCC Forecast data sheet")
    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb['VCC Forecast data']
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=19, values_only=True), 1):
        vals = [str(c)[:30] if c is not None else '-' for c in row[:14]]
        print(f"  Row{i}: {vals}")
    wb.close()
