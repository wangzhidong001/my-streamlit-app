import openpyxl
import os

base = r'C:\Users\ruijie\Desktop\IDC数据文件'
out_dir = os.path.join(base, '汇总结果')

# 1. Existing WLAN.xlsx - full header row
fpath = os.path.join(out_dir, 'WLAN.xlsx')
wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
ws = wb['WLAN']
row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print('=== WLAN.xlsx header (all columns) ===')
for i, c in enumerate(row1):
    print(f'  Col{i}: {c}')
row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
print(f'\nRow2 sample: {[str(c)[:20] if c else "-" for c in row2]}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
wb.close()

# 2. Existing IDC全产品数据.xlsx
fpath2 = os.path.join(out_dir, 'IDC全产品数据.xlsx')
wb = openpyxl.load_workbook(fpath2, read_only=True, data_only=True)
print(f'\n=== IDC全产品数据.xlsx - sheets: {wb.sheetnames} ===')
for sn in wb.sheetnames:
    ws = wb[sn]
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if row1 and row1[0]:
        print(f'  [{sn}] Header: {[str(c)[:25] if c else "-" for c in row1[0]][:25]}')
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    if row2 and row2[0]:
        print(f'  [{sn}] Row2: {[str(c)[:20] if c else "-" for c in row2[0]][:20]}')
    print(f'  [{sn}] Max row: {ws.max_row}, Max col: {ws.max_column}')
wb.close()

# 3. Switch Forecast Product - scan more rows for actual data
fpath3 = os.path.join(base, 'IDC China Quarterly Ethernet Switch Forecast, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath3, read_only=True, data_only=True)
ws = wb['Product']
print('\n=== Switch Forecast Product - scan rows 15-40 ===')
for i, row in enumerate(ws.iter_rows(min_row=15, max_row=40, values_only=True), 15):
    non_null = [(j, str(c)[:25]) for j, c in enumerate(row) if c is not None]
    if non_null:
        print(f'  Row{i}: {non_null[:10]}')
wb.close()

# 4. Switch Forecast DC Product - scan rows
ws2_sheet = wb['DC Product'] if 'DC Product' in wb.sheetnames else None
wb = openpyxl.load_workbook(fpath3, read_only=True, data_only=True)
ws = wb['DC Product']
print('\n=== Switch Forecast DC Product - scan rows 15-40 ===')
for i, row in enumerate(ws.iter_rows(min_row=15, max_row=40, values_only=True), 15):
    non_null = [(j, str(c)[:25]) for j, c in enumerate(row) if c is not None]
    if non_null:
        print(f'  Row{i}: {non_null[:10]}')
wb.close()

# 5. Forecast-Segmentation Vertical sheet
fpath4 = os.path.join(base, 'IDC China Quarterly Ethernet Switch Forecast-Segmentation, 2025Q3.xlsx')
wb = openpyxl.load_workbook(fpath4, read_only=True, data_only=True)
print(f'\n=== Switch Forecast-Seg 2025Q3 - sheets: {wb.sheetnames} ===')
ws = wb['Vertical']
print('=== Vertical sheet - scan rows 1-40 ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
    non_null = [(j, str(c)[:25]) for j, c in enumerate(row) if c is not None]
    if non_null:
        print(f'  Row{i}: {non_null[:12]}')
wb.close()

# 6. VCC Forecast data sheet - full scan
fpath5 = os.path.join(base, 'IDC_China Semiannual Virtual Client Computing Software Tracker, 2025H2.xlsx')
wb = openpyxl.load_workbook(fpath5, read_only=True, data_only=True)
ws = wb['VCC Forecast data']
print('\n=== VCC Forecast data - full scan (19 rows) ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=19, values_only=True), 1):
    non_null = [(j, str(c)[:30]) for j, c in enumerate(row) if c is not None]
    if non_null:
        print(f'  Row{i}: {non_null[:15]}')
wb.close()
