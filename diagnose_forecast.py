"""诊断预测数文件结构"""
import openpyxl
import os
from openpyxl import load_workbook

DATA_DIR = r'C:\Users\ruijie\Desktop\IDC数据文件'

# 检查WLAN Forecast文件
forecast_files = [
    'IDC China Quarterly WLAN Forecast, 2025Q4.xlsx',
    'IDC China Quarterly WLAN Forecast-Segmentation, 2025Q4.xlsx',
    'IDC China Quarterly Ethernet Switch Forecast, 2025Q4.xlsx',
]

for fname in forecast_files:
    fpath = os.path.join(DATA_DIR, fname)
    if not os.path.exists(fpath):
        print(f"NOT FOUND: {fname}")
        continue
    
    print(f"\n{'='*60}")
    print(f"文件: {fname}")
    print(f"{'='*60}")
    
    wb = load_workbook(fpath, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  --- Sheet: {sheet_name} ---")
        print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # 读取前10行
        rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
        for i, row in enumerate(rows, 1):
            vals = [str(c)[:20] if c is not None else '-' for c in row[:25]]
            print(f"  Row{i}: {vals}")
    
    wb.close()
