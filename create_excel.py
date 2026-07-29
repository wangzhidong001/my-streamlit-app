import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "阿米巴报表指标业务含义"

data = [
    ("6101", "研究与开发费", "研发部门为新产品、新技术、新工艺的研发活动所发生的各项费用，包括研发人员薪酬、研发材料、测试费用、知识产权费用等", "研发/技术"),
    ("6102", "产品资质办理费", "为获取产品所需的行业资质认证、许可证、检测报告等而支付的申请费、检测费、咨询费等", "研发/质量"),
    ("6134", "整改费", "因产品或流程不符合标准要求而进行整改所发生的费用，包括返工、修复、重新测试等", "质量/生产"),
    ("6105", "交通费", "员工因公外出办事产生的市内交通费用，如出租车费、公交费、地铁费等", "日常运营"),
    ("6106", "手机费", "员工因工作需要产生的手机通讯费用，通常按公司标准报销或补贴", "日常运营"),
    ("6103", "差旅费", "员工因公出差产生的交通（机票/高铁）、住宿、出差补贴等相关费用", "日常运营"),
    ("6104", "业务费", "日常业务活动中产生的各项杂项费用，如客户接待、业务联络等", "销售/运营"),
    ("6111", "营销费", "为推广产品或品牌而发生的广告宣传、市场推广、展会活动、宣传物料等费用", "市场/销售"),
    ("6132", "三方费用", "向第三方服务商支付的服务费用，如外包服务、技术服务、劳务派遣等", "运营/采购"),
    ("6133", "销售领料费", "销售过程中领用的样品、演示设备、宣传材料等相关物料的成本费用", "销售"),
    ("6203", "中标服务费", "参与项目投标中标后向招标平台或中介机构支付的服务费用/佣金", "销售"),
    ("6209", "项目测试费", "为特定项目进行的测试活动所发生的费用，包括测试环境搭建、第三方测试等", "项目/质量"),
    ("6217", "应收款利息", "对应收账款占用资金所计提的资金成本/利息支出", "财务"),
    ("6218", "销售借用设备占用利息", "销售部门借用设备（如样机）占用资金所产生的利息成本", "销售/财务"),
    ("6219", "测试借用设备使用费", "测试部门借用设备所产生的使用费、折旧费或租赁费", "项目/测试"),
    ("6109", "培训费", "员工参加内外部培训课程、技术交流、认证考试等所发生的培训相关费用", "人力/管理"),
    ("6110", "会议费", "组织或参加各类会议所产生的场地租赁、会议服务、食宿等费用", "管理/运营"),
    ("6135", "加班餐费", "员工因加班工作产生的餐饮费用，通常按公司加班管理制度报销", "人力/行政"),
    ("6112", "团队建设费", "部门或团队组织建设活动（团建、聚餐、户外拓展等）所发生的费用", "人力/行政"),
    ("6108", "固资折旧费", "固定资产（设备、车辆、办公家具等）按月计提的折旧费用", "财务/资产"),
    ("6113", "低值易耗费", "低值易耗品（工具、劳保用品、办公用品等）的摊销或领用成本", "行政/资产"),
    ("6321", "咨询费", "聘请外部咨询机构提供专业咨询服务所支付的费用，如管理咨询、技术咨询等", "管理"),
    ("6319", "固定电话费", "办公场所固定电话的月租费、通话费、长途费等通讯费用", "行政"),
    ("6318", "修理费", "对固定资产、办公设备、设施等进行日常维修和保养所发生的费用", "行政/资产"),
    ("6323", "审计中介费", "聘请会计师事务所、审计机构进行审计服务所支付的费用", "财务/管理"),
    ("6325", "财产保险费", "为企业财产（设备、存货、车辆等）购买保险所支付的保费", "财务/资产"),
    ("6327", "工会经费", "按工资总额的一定比例提取，用于工会活动和员工福利的经费", "人力/财务"),
    ("6322", "管理税费", "管理部门相关的税费支出，如房产税、车船使用税、印花税等", "财务"),
    ("6330", "集团征收", "集团层面向各业务单元/阿米巴收取的管理费、服务费或资源使用费", "集团管理"),
    ("6138", "NRE费用", "Non-Recurring Engineering（一次性工程费用），为客户定制化开发、模具制作等一次性投入的研发费用", "研发/项目"),
    ("6139", "动力费", "生产或办公过程中消耗的水、电、气、暖等能源动力费用", "生产/行政"),
    ("6605", "管理费用征收", "公司管理层向各业务单元征收的公共管理费用分摊", "管理/财务"),
    ("6313", "租赁物业费", "办公场所、仓库等物业的租金及物业管理费用", "行政"),
    ("6308", "招聘费", "招聘过程中产生的猎头服务费、招聘平台费用、面试差旅等费用", "人力"),
    ("6316", "办公费", "日常办公所需的各类消耗品费用，如纸张、文具、打印耗材等", "行政"),
    ("6335", "法务费", "聘请律师、法律顾问提供法律服务所支付的费用，含诉讼费、合同审查等", "管理/法务"),
    ("6214", "翻译费", "因涉外业务需要而产生的文件翻译、口译服务等相关费用", "销售/项目"),
    ("6121", "其他费用", "不属于以上各类别的其他杂项费用，用于归集难以明确分类的支出", "通用"),
    ("6114", "邮政快递费", "日常寄送信函、文件、物品等产生的邮政和快递服务费用", "行政"),
    ("6215", "运费", "产品销售过程中发生的运输、配送、物流等相关费用", "销售/物流"),
    ("6117", "服务销售成本", "提供服务类业务过程中直接发生的人工成本、材料成本和分摊费用", "销售/财务"),
    ("6312", "维修材料费", "设备维修、设施维护过程中领用的备品备件、维修材料等费用", "资产/行政"),
    ("5852", "标准服务费", "按照合同约定或服务标准收取/支付的标准化的技术服务、维护服务等费用", "服务/运营"),
    ("6126", "售前支持及专项服务费", "为项目售前阶段提供技术支持、方案设计、POC验证等专项服务所发生的费用", "销售/技术"),
    ("6129", "产品批次处理费", "对产品进行批次管理、批次追溯、批次检测等环节所发生的处理费用", "生产/质量"),
    ("6521", "自有职能费用分摊", "公司内部职能部门（人力、财务、IT等）的运营费用按规则分摊到各阿米巴单元", "管理/财务"),
    ("6525", "销售支持费用分摊", "销售支持部门（售前、方案、商务等）的费用按规则分摊到各阿米巴单元", "销售/管理"),
    ("6522", "研究平台费用分摊", "公共研究平台、共享实验室、测试平台等的运营费用按使用量分摊到各阿米巴单元", "研发/财务"),
]

headers = ["科目编码", "科目名称", "业务含义", "所属模块"]

header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="D9DEE7"),
    right=Side(style="thin", color="D9DEE7"),
    top=Side(style="thin", color="D9DEE7"),
    bottom=Side(style="thin", color="D9DEE7"),
)

zebra_fill_1 = PatternFill(fill_type="solid", fgColor="FFFFFF")
zebra_fill_2 = PatternFill(fill_type="solid", fgColor="F2F7FB")

code_font = Font(name="Consolas", size=10, color="2E75B6")
name_font = Font(name="Arial", size=10, bold=True, color="1F2937")
desc_font = Font(name="Arial", size=10, color="374151")
module_font = Font(name="Arial", size=9, color="6B7280")

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

ws.row_dimensions[1].height = 30

for row_idx, row_data in enumerate(data, 2):
    fill = zebra_fill_1 if (row_idx - 2) % 2 == 0 else zebra_fill_2
    
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        if col_idx == 1:
            cell.font = code_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            cell.font = name_font
        elif col_idx == 3:
            cell.font = desc_font
        elif col_idx == 4:
            cell.font = module_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[row_idx].height = 45

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 80
ws.column_dimensions["D"].width = 16

ws.auto_filter.ref = f"A1:D{len(data)+1}"
ws.freeze_panes = "A2"

output_path = r"C:\Users\ruijie\AppData\Roaming\TRAE SOLO CN\ModularData\ai-agent\work-mode-projects\6a4f8645e4b01f7722a32ca4\阿米巴报表指标业务含义.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
