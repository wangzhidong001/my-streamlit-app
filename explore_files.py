import openpyxl
import os

base = r'C:\Users\ruijie\Desktop\IDC数据文件'

# 1. Switch Tracker (non-seg) - 'switch' sheet
fpath = os.path.join(base, 'IDC China Quarterly Ethernet Switch Tracker, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
print('=== Switch Tracker 2025Q4 - switch sheet ===')
ws = wb['switch']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    vals = [str(c)[:25] if c is not None else '-' for c in row][:20]
    print(f'  Row{i}: {vals}')
print(f'  Max col: {ws.max_column}')
wb.close()

# 2. Switch Tracker (non-seg) - 'Product' sheet
wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
print('\n=== Switch Tracker 2025Q4 - Product sheet ===')
ws = wb['Product']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    vals = [str(c)[:25] if c is not None else '-' for c in row][:20]
    print(f'  Row{i}: {vals}')
wb.close()

# 3. Switch Forecast - Product sheet
fpath2 = os.path.join(base, 'IDC China Quarterly Ethernet Switch Forecast, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath2, read_only=True, data_only=True)
print('\n=== Switch Forecast 2025Q4 - Product sheet ===')
ws = wb['Product']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
    vals = [str(c)[:20] if c is not None else '-' for c in row][:25]
    print(f'  Row{i}: {vals}')
print(f'  Max col: {ws.max_column}, Max row: {ws.max_row}')
wb.close()

# 4. Switch Forecast - DC Product sheet
wb = openpyxl.load_workbook(fpath2, read_only=True, data_only=True)
print('\n=== Switch Forecast 2025Q4 - DC Product sheet ===')
ws = wb['DC Product']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
    vals = [str(c)[:20] if c is not None else '-' for c in row][:25]
    print(f'  Row{i}: {vals}')
wb.close()

# 5. VCC Tracker - VCC sheet and VCC Forecast data
fpath3 = os.path.join(base, 'IDC_China Semiannual Virtual Client Computing Software Tracker, 2025H2.xlsx')
wb = openpyxl.load_workbook(fpath3, read_only=True, data_only=True)
print('\n=== VCC Tracker 2025H2 - VCC sheet ===')
ws = wb['VCC']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    vals = [str(c)[:25] if c is not None else '-' for c in row][:15]
    print(f'  Row{i}: {vals}')
print(f'  Max col: {ws.max_column}, Max row: {ws.max_row}')

print('\n=== VCC Tracker 2025H2 - VCC Forecast data sheet ===')
ws = wb['VCC Forecast data']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
    vals = [str(c)[:25] if c is not None else '-' for c in row][:15]
    print(f'  Row{i}: {vals}')
print(f'  Max col: {ws.max_column}, Max row: {ws.max_row}')
wb.close()

# 6. Check WLAN Tracker structure
fpath4 = os.path.join(base, 'IDC China Quarterly WLAN Tracker, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath4, read_only=True, data_only=True)
print('\n=== WLAN Tracker 2025Q4 - sheets ===')
print(f'  Sheets: {wb.sheetnames}')
for sn in wb.sheetnames:
    ws = wb[sn]
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if row1 and row1[0]:
        first_val = row1[0][0]
        if first_val is not None:
            print(f'  [{sn}] Row1[0]={str(first_val)[:30]}')
wb.close()

# 7. Check Router Tracker structure
fpath5 = os.path.join(base, 'IDC China Quarterly Router Tracker, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath5, read_only=True, data_only=True)
print('\n=== Router Tracker 2025Q4 - sheets ===')
print(f'  Sheets: {wb.sheetnames}')
for sn in wb.sheetnames:
    ws = wb[sn]
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if row1 and row1[0]:
        first_val = row1[0][0]
        if first_val is not None:
            print(f'  [{sn}] Row1[0]={str(first_val)[:30]}')
wb.close()

# 8. Forecast-Segmentation file
fpath6 = os.path.join(base, 'IDC China Quarterly WLAN Forecast-Segmentation, 2025Q4.xlsx')
wb = openpyxl.load_workbook(fpath6, read_only=True, data_only=True)
print('\n=== WLAN Forecast-Segmentation 2025Q4 - sheets ===')
print(f'  Sheets: {wb.sheetnames}')
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    if rows and rows[0] and rows[0][0] is not None:
        print(f'  [{sn}] Row1: {[str(c)[:20] if c is not None else "-" for c in rows[0]][:15]}')
wb.close()
